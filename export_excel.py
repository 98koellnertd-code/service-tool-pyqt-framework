#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
export_excel.py
Excel-Export für Arbeitszeiten (Servicezeitenmeldung, von Grund auf neu
erzeugt) und Reisekosten (FB_0020-Vorlage wird befüllt).

Portiert aus dem alten tkinter-Tool (az_reisekosten.py), Logik unverändert
übernommen — nur an die hier vorhandenen Hilfsfunktionen (utils.py) und an
PyQt6 angepasst (keine tkinter-Variablen mehr, stattdessen Klartext-Werte
als Parameter).

openpyxl wird lazy importiert (erst beim tatsächlichen Export), damit der
App-Start nicht durch den Excel-Bibliotheksimport verzögert wird.
"""

import os
import shutil
import datetime
import calendar

from utils import (
    DAY_NAMES, MONTH_NAMES,
    calc_entry_hours, calc_day_hours, entry_detail_text, week_dates,
)


def _str_to_time(s):
    """'08:30' → datetime.time(8, 30) für Excel-Zeitwerte."""
    try:
        h, m = map(int, s.split(":"))
        return datetime.time(h, m)
    except Exception:
        return None


def default_stundenblatt_filename(month_name, year, profile):
    name_str = f"{profile.get('nachname','')}_{profile.get('vorname','')}".strip("_")
    return f"Servicezeitenmeldung_{month_name}_{year}_{name_str}.xlsx"


def default_reisekosten_filename(kw, year, profile):
    name_str = f"{profile.get('nachname','')}_{profile.get('vorname','')}".strip("_")
    return f"Reisekosten_KW{kw:02d}_{year}_{name_str}.xlsx"


# ══════════════════════════════════════════════════════════════════════════════
# Servicezeitenmeldung (Arbeitszeiten) — von Grund auf neu erzeugt
# ══════════════════════════════════════════════════════════════════════════════

def export_stundenblatt(dest, month_data, month, year, profile, wohnort="Wohnort"):
    """
    Erstellt die Servicezeitenmeldung (Monats-Excel) unter 'dest'.
    month_data: {iso_date: day_data} bereits für den Monat zusammengeführt
    (z.B. via utils.load_month_data + UI-Snapshot).
    Gibt (ad_days, id_days, total_h) zur Anzeige im Erfolgs-Dialog zurück.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    month_name = MONTH_NAMES[month - 1]

    wb = Workbook()
    ws = wb.active
    ws.title = f"{month_name} {year}"

    # ── Farben ────────────────────────────────────────────────────────────────
    CLR = {
        "header_bg": "1E3A5F", "header_fg": "FFFFFF",
        "sub_bg":    "2D5986", "sub_fg":    "FFFFFF",
        "ad":        "D6E4F0", "ad_dark":   "AED6F1",
        "id":        "E8F5E9", "id_dark":   "C8E6C9",
        "krank":     "FDECEA", "urlaub":    "FFF9C4",
        "frei":      "F5F5F5", "we":        "EEEEEE",
        "border":    "BDBDBD",
    }

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def font(bold=False, color="000000", size=10, italic=False):
        return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)

    def align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    thin = Side(style="thin", color=CLR["border"])
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Logo-Bereich / Titel-Header (Zeilen 1–5) ────────────────────────────────
    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 32
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 10

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = "Servicezeitenmeldung"
    c.font  = Font(name="Calibri", bold=True, size=18, color=CLR["header_fg"])
    c.fill  = fill(CLR["header_bg"])
    c.alignment = align("left", "center")
    for col in "BCDEFGH":
        ws[f"{col}2"].fill = fill(CLR["header_bg"])

    ws.merge_cells("I2:M2")
    c2 = ws["I2"]
    c2.value = "K&B Coding GmbH — Servicetechniker"
    c2.font  = Font(name="Calibri", bold=False, size=11, color="CCDDEE", italic=True)
    c2.fill  = fill(CLR["header_bg"])
    c2.alignment = align("right", "center")
    for col in "JKLM":
        ws[f"{col}2"].fill = fill(CLR["header_bg"])

    name    = f"{profile.get('vorname','')} {profile.get('nachname','')}".strip()
    pers_nr = profile.get("pers_nr", "")

    ws.merge_cells("A3:D3")
    ws["A3"].value     = f"  Name:  {name}"
    ws["A3"].font      = Font(name="Calibri", bold=True, size=11, color=CLR["header_fg"])
    ws["A3"].fill      = fill(CLR["sub_bg"])
    ws["A3"].alignment = align("left", "center")
    for col in "BCD":
        ws[f"{col}3"].fill = fill(CLR["sub_bg"])

    ws.merge_cells("E3:H3")
    ws["E3"].value     = f"Personal-Nr:  {pers_nr}"
    ws["E3"].font      = Font(name="Calibri", bold=True, size=11, color=CLR["header_fg"])
    ws["E3"].fill      = fill(CLR["sub_bg"])
    ws["E3"].alignment = align("center", "center")
    for col in "FGH":
        ws[f"{col}3"].fill = fill(CLR["sub_bg"])

    ws.merge_cells("I3:O3")
    ws["I3"].value     = f"Monat:  {month_name} {year}"
    ws["I3"].font      = Font(name="Calibri", bold=True, size=11, color=CLR["header_fg"])
    ws["I3"].fill      = fill(CLR["sub_bg"])
    ws["I3"].alignment = align("right", "center")
    for col in ["J", "K", "L", "M", "N", "O"]:
        ws[f"{col}3"].fill = fill(CLR["sub_bg"])

    ws.merge_cells("A4:O4")
    ws["A4"].value = (
        "  🔵 Außendienst (AD)     🟢 Innendienst / Home Office (ID/HO)"
        "     🔴 Krank     🟡 Urlaub / Sonderurlaub     ⬜ Wochenende / Frei")
    ws["A4"].font      = Font(name="Calibri", size=9, color="555555", italic=True)
    ws["A4"].alignment = align("left", "center")

    # ── Spaltenköpfe (Zeile 6) ───────────────────────────────────────────────
    col_widths2 = {
        "A": 5, "B": 12, "C": 5, "D": 18, "E": 8, "F": 8,
        "G": 7, "H": 9, "I": 9, "J": 10,
        "K": 5, "L": 14, "M": 10, "N": 10, "O": 22,
    }
    for col, w in col_widths2.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[6].height = 30
    headers = [
        "Nr", "Datum", "Wt", "Dienstart", "Start", "Ende",
        "Pause\n(min)", "Netto\n(h)", "Soll\n(h)", "Gleitzeit\n(+/−)",
        "Eintrag\n#", "Auftrag-Nr", "Kundenname", "Standort", "Details",
    ]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=ci, value=h)
        cell.font      = Font(name="Calibri", bold=True, size=10, color=CLR["header_fg"])
        cell.fill      = fill(CLR["header_bg"])
        cell.alignment = align("center", "center", wrap=True)
        cell.border    = Border(
            left=Side(style="medium", color="FFFFFF"),
            right=Side(style="medium", color="FFFFFF"),
            bottom=Side(style="medium", color="5B8DB8"))
    ws.cell(6, 10).fill = fill("7B5EA7")

    ws.auto_filter.ref = "A6:O6"
    ws.freeze_panes    = "A7"

    # ── Daten (Zeilen 7+) ────────────────────────────────────────────────────
    SOLL_H   = 8.0
    cal_days = calendar.monthrange(year, month)[1]
    data_row = 7
    ad_days  = 0
    id_days  = 0
    total_h  = 0.0
    gleit_h  = 0.0

    for dn in range(1, cal_days + 1):
        d     = datetime.date(year, month, dn)
        ds    = d.isoformat()
        day   = month_data.get(ds, {})
        wd    = DAY_NAMES[d.weekday()]
        is_we = d.weekday() >= 5

        status  = day.get("status", "Frei" if is_we else "Arbeit")
        entries = day.get("entries") or [{}]

        netto = calc_day_hours(day)

        is_arbeits = status == "Arbeit" and not is_we
        soll  = SOLL_H if is_arbeits else 0.0
        gleit = round(netto - soll, 2) if is_arbeits else 0.0

        if status == "Krank":
            day_art  = "Krank"
            day_fill = fill(CLR["krank"])
            soll = SOLL_H; gleit = 0.0
        elif status in ("Urlaub", "Feiertag", "Sonstiges"):
            day_art  = status
            day_fill = fill(CLR["urlaub"])
            soll = SOLL_H; gleit = 0.0
        elif status in ("GLZ", "Kurzarbeit"):
            day_art  = f"{status} (−8h)"
            day_fill = fill(CLR["urlaub"])
            soll = SOLL_H; gleit = -SOLL_H
        elif is_we or status == "Frei":
            day_art  = ""
            day_fill = fill(CLR["we"])
        else:
            day_art  = None
            day_fill = None

        total_h += netto
        gleit_h  = round(gleit_h + gleit, 2)

        gleit_str = (f"+{gleit:.2f}" if gleit > 0
                     else f"{gleit:.2f}" if gleit < 0 else "0")
        gleit_color = ("006100" if gleit > 0 else "9C0006" if gleit < 0 else "555555")

        def _entry_fill(e, ei):
            if day_fill is not None:
                return day_art, day_fill
            ed = e.get("dienst", "Außendienst")
            if ed in ("Innendienst", "Homeoffice", "Home Office"):
                return "Innendienst / HO", fill(CLR["id"] if ei % 2 == 0 else CLR["id_dark"])
            elif ed == "Ausland":
                return "Ausland (AD)", fill("D5E8D4")
            else:
                return "Außendienst", fill(CLR["ad"] if ei % 2 == 0 else CLR["ad_dark"])

        for e in entries:
            ed = e.get("dienst", "Außendienst")
            if day_fill is None:
                if ed in ("Innendienst", "Homeoffice", "Home Office"):
                    id_days += 1
                else:
                    ad_days += 1
        if len(entries) > 1:
            ad_days -= (len(entries) - 1)

        start = day.get("start", "")
        end   = day.get("end", "")
        pause = day.get("pause", "")

        for ei, entry in enumerate(entries):
            ws.row_dimensions[data_row].height = 18
            art_text, row_fill = _entry_fill(entry, ei)

            route = entry_detail_text(entry, wohnort)

            e_start = entry.get("start") or (start if ei == 0 else "")
            e_end   = entry.get("end")   or (end   if ei == 0 else "")
            e_pause = entry.get("pause") or (pause if ei == 0 else "")
            e_netto = calc_entry_hours(entry) if entry.get("start") and entry.get("end") \
                      else (netto if ei == 0 else 0.0)

            if ei == 0:
                day_vals = [dn, d.strftime("%d.%m.%Y"), wd, art_text,
                            e_start, e_end,
                            int(e_pause) if e_pause and str(e_pause).isdigit() else "",
                            f"{e_netto:.2f}" if e_netto else "",
                            f"{soll:.1f}" if soll else "",
                            gleit_str]
            else:
                day_vals = ["", d.strftime("%d.%m.%Y"), "", art_text,
                            e_start, e_end,
                            int(e_pause) if e_pause and str(e_pause).isdigit() else "",
                            f"{e_netto:.2f}" if e_netto else "",
                            "", ""]

            row_vals = day_vals + [
                ei + 1,
                entry.get("auftr_nr", ""),
                entry.get("kunde_name", ""),
                entry.get("standort", ""),
                route,
            ]

            for ci, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=data_row, column=ci, value=val)
                cell.fill   = row_fill
                cell.border = border_all
                cell.font   = Font(name="Calibri", size=10, bold=(ci == 4),
                                    color="333333" if not is_we else "888888")
                wrap = (ci == 15)
                cell.alignment = align(
                    "center" if ci in (1, 3, 5, 6, 7, 8, 9, 11) else "left", wrap=wrap)

            if ei == 0 and gleit != 0:
                ws.cell(data_row, 10).font = Font(name="Calibri", bold=True, size=10,
                                                    color=gleit_color)
            if ei > 0:
                ws.cell(data_row, 11).fill = fill("E3F2FD")

            data_row += 1

    max_detail_len = 30
    for row_cells in ws.iter_rows(min_row=7, max_row=data_row - 1, min_col=15, max_col=15):
        for cell in row_cells:
            if cell.value:
                max_detail_len = max(max_detail_len, len(str(cell.value)) + 4)
    ws.column_dimensions["O"].width = min(max_detail_len, 80)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows       = "6:6"
    ws.oddHeader.center.text  = f"Servicezeitenmeldung — {name} — {month_name} {year}"
    ws.oddFooter.right.text   = "Seite &P von &N"

    wb.save(dest)
    return ad_days, id_days, total_h


# ══════════════════════════════════════════════════════════════════════════════
# Reisekostenabrechnung (FB_0020-Vorlage befüllen)
# ══════════════════════════════════════════════════════════════════════════════

def _ws_set(ws, row, col, value, fmt=None):
    """Schreibt in eine Zelle — überspringt MergedCell (non-top-left) sicher."""
    from openpyxl.cell.cell import MergedCell
    try:
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            return
        cell.value = value
        if fmt:
            cell.number_format = fmt
    except Exception:
        pass


def _entry_is_filled(entry):
    """
    Ein Eintrag gilt nur dann als 'befüllt', wenn er mindestens ein
    inhaltliches Feld trägt (Start/Ende-Zeit oder Auftrags-/Kundendaten).
    Unbearbeitete Tage liefern sonst nur den leeren Platzhalter-Eintrag
    "[{}]" — der soll in der Reisekostenabrechnung nicht als Zeile auftauchen.
    """
    return bool(
        entry.get("start") or entry.get("end") or entry.get("kunde_name")
        or entry.get("standort") or entry.get("auftr_nr"))


def export_reisekosten(tmpl, dest, kw_data, kw, year, profile, wohnort="Wohnort"):
    """
    Kopiert die FB_0020-Vorlage 'tmpl' nach 'dest' und befüllt sie mit den
    Daten der angegebenen Kalenderwoche. kw_data ist das KW-JSON-Dict
    (inkl. optionalem "_extras" für Sonstiges). Gibt die Anzahl befüllter
    Reise-Blöcke zurück.
    """
    import openpyxl

    shutil.copy(tmpl, dest)
    wb = openpyxl.load_workbook(dest, keep_vba=False)
    ws = wb.active

    name  = f"{profile.get('nachname','')}, {profile.get('vorname','')}".strip(", ")
    dates = week_dates(year, kw)

    _ws_set(ws, 4, 2, name)
    _ws_set(ws, 5, 2, "K&B Coding - Service")
    _ws_set(ws, 7, 2, "Service")
    _ws_set(ws, 7, 6, dates[0], "DD.MM.YY")
    _ws_set(ws, 7, 8, dates[-1], "DD.MM.YY")

    kw_extras = kw_data.get("_extras", {})

    block = 0
    for ds in sorted(k for k in kw_data if k != "_extras"):
        day = kw_data[ds]
        # Unbefüllte Tage (z.B. Wochenende oder einfach nie bearbeitet)
        # komplett überspringen — die tauchen sonst als leere Zeile auf.
        if not day:
            continue
        try:
            d = datetime.date.fromisoformat(ds)
        except Exception:
            continue
        default_status = "Frei" if d.weekday() >= 5 else "Arbeit"
        if day.get("status", default_status) != "Arbeit":
            continue
        entries  = day.get("entries") or [{}]
        first_ad = True
        for entry in entries:
            dienst = entry.get("dienst", "Außendienst")
            # Reisekosten gibt es nur für Außendienst — Innendienst-Einträge
            # (inkl. deren Allgemeinkosten-Details) gehören nur in die
            # Monatsübersicht/Servicezeitenmeldung, nicht hierher.
            if dienst in ("Innendienst", "Homeoffice", "Home Office"):
                continue
            # Leere Platzhalter-Einträge (unbearbeiteter Tag) ebenfalls
            # nicht exportieren.
            if not _entry_is_filled(entry):
                continue

            ra = 12 + block * 3
            rb = ra + 1
            rc = ra + 2
            if ra > 188:
                break

            e_start = entry.get("start") or (day.get("start", "") if first_ad else "")
            e_end   = entry.get("end")   or (day.get("end", "")   if first_ad else "")
            land    = entry.get("land")  or day.get("land", "DE") or "DE"
            reise   = entry_detail_text(entry, wohnort)
            t_s = _str_to_time(e_start)
            t_e = _str_to_time(e_end)

            if first_ad:
                _ws_set(ws, ra, 1, d, "DD.MM.YY")
            if t_s:
                _ws_set(ws, ra, 2, t_s, "HH:MM")
            if reise:
                _ws_set(ws, ra, 4, reise)
            if first_ad and day.get("uebernacht") == "ja":
                _ws_set(ws, ra, 9, "ja")
            if first_ad and day.get("fruehstueck") == "ja":
                _ws_set(ws, ra, 10, "ja")

            if t_e:
                _ws_set(ws, rb, 2, t_e, "HH:MM")
            _ws_set(ws, rb, 6, land)
            if first_ad and day.get("mittag") == "ja":
                _ws_set(ws, rb, 10, "ja")

            if first_ad and day.get("abend") == "ja":
                _ws_set(ws, rc, 10, "ja")

            first_ad = False
            block += 1

    try:
        s_txt = kw_extras.get("sonstiges_txt", "").strip()
        s_val = float(kw_extras.get("sonstiges", "") or 0)
        if s_val or s_txt:
            _ws_set(ws, 195, 7, s_txt or "Sonstige Kosten")
            if s_val:
                _ws_set(ws, 195, 16, s_val)
    except Exception:
        pass

    wb.save(dest)
    return block


# ══════════════════════════════════════════════════════════════════════════════
# Bestellung (Ersatzteile) — von Grund auf neu erzeugt
# ══════════════════════════════════════════════════════════════════════════════

def default_bestellung_filename(bestellung):
    """Dateiname-Vorschlag, z.B. 'Bestellung_3_Max_Mustermann.xlsx'."""
    name_str = (bestellung.get("name") or "Bestellung").strip().replace(" ", "_")
    return f"Bestellung_{bestellung.get('id', '')}_{name_str}.xlsx"


def export_bestellung(dest, bestellung):
    """
    Erstellt eine Excel-Bestellung unter 'dest'. 'bestellung' ist das
    Bestellungs-dict ({"id", "name", "datum", "teile": [...]}) wie es auch
    als JSON in res/orders/<id>.json liegt. Jedes Element in "teile" hat
    die Felder order_no, name und anzahl.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Bestellung"

    header_bg, header_fg = "1E3A5F", "FFFFFF"

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def align(h="left", v="center"):
        return Alignment(horizontal=h, vertical=v)

    thin = Side(style="thin", color="BDBDBD")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Kopfbereich: Titel, Name, Datum ────────────────────────────────────
    ws.merge_cells("A1:C1")
    ws["A1"].value     = "Ersatzteilbestellung"
    ws["A1"].font      = Font(name="Calibri", bold=True, size=16, color=header_fg)
    ws["A1"].fill      = fill(header_bg)
    ws["A1"].alignment = align("left")
    ws.row_dimensions[1].height = 30

    ws["A2"].value = "Name:"
    ws["A2"].font  = Font(name="Calibri", bold=True, size=10)
    ws["B2"].value = bestellung.get("name", "")

    ws["A3"].value = "Datum:"
    ws["A3"].font  = Font(name="Calibri", bold=True, size=10)
    ws["B3"].value = bestellung.get("datum", "")

    bestell_id = bestellung.get("id")
    if bestell_id is not None:
        ws["A4"].value = "Bestell-Nr. (intern):"
        ws["A4"].font  = Font(name="Calibri", bold=True, size=10)
        ws["B4"].value = bestell_id

    # ── Spaltenköpfe der Positionsliste ────────────────────────────────────
    col_widths = {"A": 20, "B": 50, "C": 12}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    head_row = 6
    headers = ["Bestellnummer", "Name", "Anzahl"]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=head_row, column=ci, value=h)
        cell.font      = Font(name="Calibri", bold=True, size=10, color=header_fg)
        cell.fill      = fill(header_bg)
        cell.alignment = align("center")
        cell.border    = border_all
    ws.auto_filter.ref = f"A{head_row}:C{head_row}"
    ws.freeze_panes    = f"A{head_row + 1}"

    # ── Positionen ──────────────────────────────────────────────────────────
    row = head_row + 1
    for teil in bestellung.get("teile", []):
        werte = [teil.get("order_no", ""), teil.get("name", ""), teil.get("anzahl", 1)]
        for ci, val in enumerate(werte, start=1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.border    = border_all
            cell.font      = Font(name="Calibri", size=10)
            cell.alignment = align("center" if ci == 3 else "left")
        row += 1

    ws.page_setup.orientation = "portrait"
    ws.oddHeader.center.text  = "Ersatzteilbestellung"
    ws.oddFooter.right.text   = "Seite &P von &N"

    wb.save(dest)
